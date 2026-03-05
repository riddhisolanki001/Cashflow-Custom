import frappe

import time

from io import BytesIO

from openpyxl import load_workbook

from openpyxl.styles import Font

from frappe.utils.xlsxutils import make_xlsx
 
 
@frappe.whitelist()

def export():
 
    max_retries = 5

    retry_delay = 1

    result = None
 
    # Retry API call

    for attempt in range(max_retries):

        result = frappe.call(

            "general_ledger_customizations.api.cash_flow_data.get_cash_flow_prepared_data"

        )
 
        if result and result.get("success"):

            break
 
        time.sleep(retry_delay)
    
    if not result or not result.get("success"):

        frappe.throw(result.get("message") if result else "Unable to generate report.")
 
    columns = result["columns"]

    data = result["data"]
 
    rows = []
 
    # Header row

    rows.append([col.get("label") for col in columns])
 
    fieldnames = [col.get("fieldname") for col in columns]
 
    # Data rows

    for row in data:

        rows.append([row.get(field) for field in fieldnames])
 
    # Create XLSX

    xlsx_file = make_xlsx(rows, "Cash Flow")
 
    wb = load_workbook(BytesIO(xlsx_file.getvalue()))

    ws = wb.active
 
    bold = Font(bold=True)

    normal = Font(bold=False)
 
    # Header Bold

    for col in range(1, ws.max_column + 1):

        ws.cell(row=1, column=col).font = bold
 
    # Row Formatting

    for excel_row_index, row_data in enumerate(data, start=2):
 
        # skip empty rows
        if not row_data:
            continue
    
        label = (
            str(row_data.get("section"))
            or str(row_data.get("section_name"))
            or ""
        ).replace("'", "").strip()
    
        indent = int(row_data.get("indent") or 0)
    
        # detect section header
        is_section_header = indent == 0 and not row_data.get("is_section_total")
    
        # detect section totals
        is_section_total = row_data.get("is_section_total") is True
    
        # detect summary rows
        summary_keywords = [
            "net increase",
            "opening cash",
            "closing cash",
            "net cash",
        ]
    
        is_summary = any(k in label.lower() for k in summary_keywords)
    
        # apply indentation
        ws.cell(row=excel_row_index, column=1).value = "   " * indent + label
    
        # decide font
        row_font = bold if (is_section_header or is_section_total or is_summary) else normal
    
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row_index, column=col).font = row_font
 
    # Freeze header row

    ws.freeze_panes = "A2"
 
    # Auto column width

    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter
 
        for cell in column:

            if cell.value:

                max_length = max(max_length, len(str(cell.value)))
 
        ws.column_dimensions[column_letter].width = max_length + 2
 
    # Save output

    output = BytesIO()

    wb.save(output)

    output.seek(0)
 
    frappe.response["filename"] = "Cash Flow.xlsx"

    frappe.response["filecontent"] = output.read()

    frappe.response["type"] = "binary"
 