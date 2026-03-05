import frappe

import time

from io import BytesIO

from openpyxl import load_workbook

from openpyxl.styles import Font

from frappe.utils.xlsxutils import make_xlsx
 
 
@frappe.whitelist()

def export():
 
    max_retries = 5

    retry_delay = 1

    result = None
 
    # Retry API call

    for attempt in range(max_retries):

        result = frappe.call(

            "general_ledger_customizations.api.profit_and_loss_statement_data.get_pl_statement_prepared_data"

        )
 
        if result and result.get("success"):

            break
 
        time.sleep(retry_delay)
 
    if not result or not result.get("success"):

        frappe.throw(result.get("message") if result else "Unable to generate report.")
 
    columns = result["columns"]

    data = result["data"]
 
    rows = []
 
    # Header row

    rows.append([col.get("label") for col in columns])
 
    fieldnames = [col.get("fieldname") for col in columns]
 
    # Data rows

    for row in data:

        rows.append([row.get(field) for field in fieldnames])
 
    # Create XLSX

    xlsx_file = make_xlsx(rows, "Profit and Loss Statement")
 
    wb = load_workbook(BytesIO(xlsx_file.getvalue()))

    ws = wb.active
 
    bold = Font(bold=True)

    normal = Font(bold=False)
 
    # Header Bold

    for col in range(1, ws.max_column + 1):

        ws.cell(row=1, column=col).font = bold
 
    # Row Formatting

    for excel_row_index, row_data in enumerate(data, start=2):
 
        account_name = str(row_data.get("account") or "").strip()
        indent = int(row_data.get("indent") or 0)
    
        # check next row indent
        next_indent = None
        if excel_row_index - 1 < len(data):
            next_row = data[excel_row_index - 1]
            next_indent = int(next_row.get("indent") or 0)
    
        # parent detection
        is_parent = next_indent is not None and next_indent > indent
    
        # detect totals
        is_total = "total" in account_name.lower()
    
        # apply indentation
        ws.cell(row=excel_row_index, column=1).value = "   " * indent + account_name
    
        # font decision
        row_font = bold if (is_parent or is_total) else normal
    
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row_index, column=col).font = row_font
 
    # Freeze header row

    ws.freeze_panes = "A2"
 
    # Auto column width

    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter
 
        for cell in column:

            if cell.value:

                max_length = max(max_length, len(str(cell.value)))
 
        ws.column_dimensions[column_letter].width = max_length + 2
 
    # Save output

    output = BytesIO()

    wb.save(output)

    output.seek(0)
 
    frappe.response["filename"] = "Profit and Loss Statement.xlsx"

    frappe.response["filecontent"] = output.read()

    frappe.response["type"] = "binary"
 